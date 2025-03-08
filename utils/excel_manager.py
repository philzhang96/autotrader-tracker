import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def manage_excel_data(input_file="car_info.xlsx"):
    """
    Manages the data in the Excel file by converting prices to numbers,
    comparing the latest prices with previous ones, and applying conditional formatting.
    
    Args:
        input_file (str): The path to the Excel file containing car price data.
    """
    print(f"Applying data management to {input_file}...")

    # Load the Excel file with pandas, ensure all column names are strings
    df = pd.read_excel(input_file, dtype=str)
    df.columns = df.columns.map(str)  # Convert all column names to strings

    # Standardize date columns by ensuring they are in "DD-MM-YYYY" format
    standardized_columns = []
    column_mapping = {}  # To handle duplicate columns
    for col in df.columns:
        try:
            # Check for known date formats and convert accordingly
            if " " in col:  # Likely a datetime format with time
                parsed_date = pd.to_datetime(col, format='%Y-%m-%d %H:%M:%S', errors='coerce')
            else:
                parsed_date = pd.to_datetime(col, dayfirst=True, errors='coerce')
            
            # Reformat the date if valid
            if parsed_date and not pd.isnull(parsed_date):
                formatted_date = parsed_date.strftime('%d-%m-%Y')
                
                # Handle duplicate column names by keeping only the latest
                if formatted_date in column_mapping:
                    print(f"Duplicate column found: {col}. Keeping the latest.")
                column_mapping[formatted_date] = col
                
                col = formatted_date
                
        except Exception as e:
            print(f"Error parsing column {col}: {e}")
        
        standardized_columns.append(col)
    
    # Apply the standardized column names to the DataFrame, keeping only the latest in case of duplicates
    df = df.loc[:, ~df.columns.duplicated(keep='last')]
    df.columns = [column_mapping.get(col, col) for col in df.columns]
    
    print(f"Standardized price columns: {df.columns}")

    # Identify columns that match the "DD-MM-YYYY" format
    price_columns = [col for col in df.columns if col.count("-") == 2 and len(col) == 10]
    print(f"Identified price columns: {price_columns}")
    
    # Convert prices to numeric values, removing "£" and handling "SOLD" as NaN
    for col in price_columns:
        df[col] = pd.to_numeric(df[col].replace('[£,]', '', regex=True), errors='coerce')
    
    # Load the workbook and sheet using openpyxl
    wb = load_workbook(input_file)
    ws = wb.active
    
    # Define color fills for conditional formatting
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    full_red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Full red for "SOLD"

    # Apply conditional formatting
    for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        # Check if the row is "SOLD" and apply full red highlight
        row_values = [cell.value for cell in row]
        
        if "SOLD" in row_values:
            for cell in row:
                cell.fill = full_red_fill
            print(f"Row {row_index} highlighted red for 'SOLD'.")
            continue
        
        # Compare the most recent two price columns if available
        if len(price_columns) >= 2:
            latest_price = df[price_columns[-1]].iloc[row_index - 2]
            previous_price = df[price_columns[-2]].iloc[row_index - 2]
            
            # Apply red if price increased, green if decreased, and no change otherwise
            if pd.notna(latest_price) and pd.notna(previous_price):
                cell = row[-1]  # The latest price cell should be the last column in the row
                if latest_price > previous_price:
                    cell.fill = red_fill
                    print(f"Price increased at row {row_index}. Highlighted red.")
                elif latest_price < previous_price:
                    cell.fill = green_fill
                    print(f"Price decreased at row {row_index}. Highlighted green.")

    # Save the updated Excel file
    print("Saving changes to Excel file...")
    try:
        wb.save(input_file)
        print(f"Data management and formatting applied to {input_file}")
    except Exception as e:
        print(f"Failed to save Excel file: {e}")
