import os
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Output Excel file to save scraped data
OUTPUT_EXCEL_FILE = r"E:\Coding Projects\autotrader_data.xlsx"
# Input Excel file containing the URLs to scrape
INPUT_EXCEL_FILE = r"E:\Coding Projects\urls_input.xlsx"

def autotrader_scraper_selenium(urls):
    # Set up Selenium WebDriver
    options = Options()
    options.add_argument("_tt_enable_cookie=1") 
    options.add_argument("--disable-blink-features=AutomationControlled")  # Helps bypass bot detection
    options.add_argument("--window-size=1920,1080")  # Set a normal window size to mimic human behavior

    driver_service = Service(r"C:\Windows\chromedriver-win64\chromedriver.exe")  # Update path to ChromeDriver
    driver = webdriver.Chrome(service=driver_service, options=options)

    results = []
    today_date = datetime.today().strftime('%d-%m-%Y')  # Get today's date

    for url in urls:
        try:
            print(f"Scraping: {url}")
            driver.get(url)

            # Check if the advert is no longer available
            try:
                not_available_element = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'The advert you are looking for is no longer available')]"))
                )
                not_available = True
            except:
                not_available = False

            # Extract price
            try:
                price_element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//h2[@data-testid='advert-price']"))
                )
                price_text = price_element.text.strip()
            except:
                price_text = "Price not found"

            # Extract make of the car
            try:
                make_element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//h1[@data-testid='advert-title']"))
                )
                make_text = make_element.text.strip()
            except:
                make_text = "Make not found"

            # Mark status
            status = "Not Available" if not_available else "Available"

            # Store the result
            results.append({"URL": url, "Make": make_text, today_date: price_text, "Status": status})

        except Exception as e:
            print(f"Error scraping {url}: {e}")
            results.append({"URL": url, "Make": "Error fetching make", today_date: "Error fetching price", "Status": "Error"})

    driver.quit()

    # Convert results to DataFrame
    df_new = pd.DataFrame(results)

    # Check if the output Excel file already exists
    if os.path.exists(OUTPUT_EXCEL_FILE):
        # Load existing data
        df_existing = pd.read_excel(OUTPUT_EXCEL_FILE, dtype=str)

        # Ensure column names match expectations
        column_mapping = {
            "url": "URL",
            "make": "Make",
            "price": today_date  # If there is an old "price" column, replace it with today's date
        }
        df_existing.rename(columns={col: column_mapping[col] for col in df_existing.columns if col in column_mapping}, inplace=True)

        # Merge new data with existing data
        df_combined = df_existing.merge(df_new, on=["URL", "Make"], how="outer")
    else:
        df_combined = df_new

    # Save to Excel
    df_combined.to_excel(OUTPUT_EXCEL_FILE, index=False)

    # Apply conditional formatting
    apply_red_highlight(OUTPUT_EXCEL_FILE)

    print(f"\n✅ Data saved to {OUTPUT_EXCEL_FILE}")

    # Print results
    for result in results:
        print(f"URL: {result['URL']}\nMake: {result['Make']}\nPrice ({today_date}): {result[today_date]}\nStatus: {result['Status']}\n")

def apply_red_highlight(filename):
    """Applies red highlight to rows where the Status is 'Not Available'."""
    wb = load_workbook(filename)
    ws = wb.active

    # Identify the "Status" column index
    status_col = None
    for col_index, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        if col[0].value == "Status":
            status_col = col_index
            break

    if status_col is None:
        print("⚠️ 'Status' column not found in the Excel file.")
        return

    # Define red fill style
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Apply formatting to rows where "Status" is 'Not Available'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=status_col, max_col=status_col):
        if row[0].value == "Not Available":
            for cell in ws[row[0].row]:
                cell.fill = red_fill

    wb.save(filename)
    wb.close()
    print("✅ Red highlighting applied for unavailable adverts.")

if __name__ == '__main__':
    # Read the input Excel file containing the URLs.
    df_input = pd.read_excel(INPUT_EXCEL_FILE)
    urls = df_input["URL"].tolist()
    autotrader_scraper_selenium(urls)
