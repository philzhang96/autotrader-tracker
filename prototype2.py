import os
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Output and Input Excel files
OUTPUT_EXCEL_FILE = r"E:\Coding Projects\autotrader_data.xlsx"
INPUT_EXCEL_FILE = r"E:\Coding Projects\urls_input.xlsx"

def autotrader_scraper_selenium(urls):
    # Set up Selenium WebDriver
    options = Options()
    options.add_argument("_tt_enable_cookie=1") 
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--window-size=1920,1080")

    driver_service = Service(r"C:\Windows\chromedriver-win64\chromedriver.exe")  
    driver = webdriver.Chrome(service=driver_service, options=options)

    results = []
    today_date = datetime.today().strftime('%d-%m-%Y')  
    unavailable_urls = []

    for url in urls:
        try:
            print(f"Scraping: {url}")
            driver.get(url)

            # Check if the advert is no longer available
            try:
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'The advert you are looking for is no longer available')]"))
                )
                unavailable_urls.append(url)
                price_text = "SOLD"  # Replace price with SOLD
                mileage_text = "N/A"
            except:
                # Extract price if available
                try:
                    price_element = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "//h2[@data-testid='advert-price']"))
                    )
                    price_text = price_element.text.strip()
                except:
                    price_text = "Price not found"

                # Extract mileage
                try:
                    mileage_element = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "//li[contains(@class, 'at__sc-1n64n0d-9') and contains(@class, 'at__sc-1ebejir-1')]"))
                    )
                    mileage_text = mileage_element.text.strip()
                except:
                    mileage_text = "Mileage not found"

            # Extract make
            try:
                make_element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//h1[@data-testid='advert-title']"))
                )
                make_text = make_element.text.strip()
            except:
                make_text = "Make not found"

            # Store the result
            results.append({"URL": url, "Make": make_text, "Mileage": mileage_text, today_date: price_text})

        except Exception as e:
            print(f"Error scraping {url}: {e}")
            results.append({"URL": url, "Make": "Error fetching make", "Mileage": "Error fetching mileage", today_date: "Error fetching price"})

    driver.quit()

    # Convert results to DataFrame
    df_new = pd.DataFrame(results)

    # Load existing data or create a new one
    if os.path.exists(OUTPUT_EXCEL_FILE):
        df_existing = pd.read_excel(OUTPUT_EXCEL_FILE, dtype=str)

        # Ensure only one Mileage column by updating values rather than creating a new one
        if "Mileage" in df_existing.columns:
            df_existing = df_existing.drop(columns=["Mileage"])  # Remove old mileage column to prevent duplicates
        
        df_combined = df_existing.merge(df_new, on=["URL", "Make"], how="outer")
    else:
        df_combined = df_new

    # Reorder columns to ensure 'Mileage' is next to 'Make'
    columns_order = ["URL", "Make", "Mileage"] + [col for col in df_combined.columns if col not in ["URL", "Make", "Mileage"]]
    df_combined = df_combined[columns_order]

    # Save updated data to Excel
    df_combined.to_excel(OUTPUT_EXCEL_FILE, index=False)

    # Apply conditional formatting for SOLD listings
    apply_red_highlight(OUTPUT_EXCEL_FILE, today_date)

    # Remove unavailable URLs from input list
    if unavailable_urls:
        remove_unavailable_urls(INPUT_EXCEL_FILE, unavailable_urls)

    print(f"\n✅ Data saved to {OUTPUT_EXCEL_FILE}")

    # Print results in terminal
    for result in results:
        print(f"URL: {result['URL']}\nMake: {result['Make']}\nMileage: {result['Mileage']}\nPrice ({today_date}): {result[today_date]}\n")

def apply_red_highlight(filename, price_column):
    """Applies red highlight to rows where the price is 'SOLD'."""
    wb = load_workbook(filename)
    ws = wb.active

    # Identify the column index for today's price
    price_col_index = None
    for col_index, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        if col[0].value == price_column:
            price_col_index = col_index
            break

    if price_col_index is None:
        print(f"⚠️ Column '{price_column}' not found in the Excel file.")
        return

    # Define red fill style
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Apply formatting to rows where the price is 'SOLD'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=price_col_index, max_col=price_col_index):
        if row[0].value == "SOLD":
            for cell in ws[row[0].row]:
                cell.fill = red_fill

    wb.save(filename)
    wb.close()
    print("✅ Red highlighting applied for 'SOLD' vehicles.")

def remove_unavailable_urls(input_filename, unavailable_urls):
    """Removes unavailable URLs from the input file."""
    if not os.path.exists(input_filename):
        print("⚠️ Input file does not exist.")
        return

    df_input = pd.read_excel(input_filename, dtype=str)

    if "URL" not in df_input.columns:
        print("⚠️ 'URL' column not found in input file.")
        return

    initial_count = len(df_input)

    # Remove rows where the URL is in the unavailable list
    df_filtered = df_input[~df_input["URL"].isin(unavailable_urls)]

    removed_count = initial_count - len(df_filtered)
    if removed_count > 0:
        df_filtered.to_excel(input_filename, index=False)
        print(f"✅ Removed {removed_count} unavailable URLs from {input_filename}")
    else:
        print("✅ No URLs needed to be removed.")

if __name__ == '__main__':
    # Read the input Excel file containing the URLs.
    df_input = pd.read_excel(INPUT_EXCEL_FILE)
    urls = df_input["URL"].tolist()
    autotrader_scraper_selenium(urls)
