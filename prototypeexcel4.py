import os
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Output and Input Excel files
OUTPUT_EXCEL_FILE = r"E:\Coding Projects\autotrader_data.xlsx"
INPUT_EXCEL_FILE = r"E:\Coding Projects\urls_input.xlsx"

def autotrader_scraper_selenium(urls):
    # Set up Selenium WebDriver
    options = Options()
    options.add_argument("_tt_enable_cookie=1") 
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--window-size=1920,1080")

    driver_service = Service(r"C:\Windows\chromedriver-win64\chromedriver.exe")  
    driver = webdriver.Chrome(service=driver_service, options=options)

    today_date = datetime.today().strftime('%d-%m-%Y')  
    unavailable_urls = []

    # Load existing data
    if os.path.exists(OUTPUT_EXCEL_FILE):
        df_existing = pd.read_excel(OUTPUT_EXCEL_FILE, dtype=str)
    else:
        df_existing = pd.DataFrame(columns=["URL", "Make", today_date])  # Create empty if not exist

    for url in urls:
        try:
            print(f"Scraping: {url}")
            driver.get(url)

            # Check if the advert is no longer available
            try:
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'The advert you are looking for is no longer available')]"))
                )
                unavailable_urls.append(url)
                price_text = "SOLD"  # Replace price with SOLD
            except:
                # Extract price if available
                try:
                    price_element = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "//h2[@data-testid='advert-price']"))
                    )
                    price_text = price_element.text.strip()
                except:
                    price_text = "Price not found"

            # Extract make
            try:
                make_element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//h1[@data-testid='advert-title']"))
                )
                make_text = make_element.text.strip()
            except:
                make_text = "Make not found"

            # Update existing DataFrame
            df_existing.loc[df_existing["URL"] == url, today_date] = price_text

        except Exception as e:
            print(f"Error scraping {url}: {e}")

    driver.quit()

    # Save updated data to Excel
    df_existing.to_excel(OUTPUT_EXCEL_FILE, index=False)

    # Apply conditional formatting (red highlight) for "SOLD" entries
    apply_red_highlight(OUTPUT_EXCEL_FILE, today_date)

    # Remove unavailable URLs from input list
    if unavailable_urls:
        remove_unavailable_urls(INPUT_EXCEL_FILE, unavailable_urls)

    print(f"\n✅ Data saved to {OUTPUT_EXCEL_FILE}")

def apply_red_highlight(filename, price_column):
    """Applies red highlight to rows where the price is 'SOLD'."""
    wb = load_workbook(filename)
    ws = wb.active

    # Identify the column index for today's price
    price_col_index = None
    for col_index, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        if col[0].value == price_column:
            price_col_index = col_index
            break

    if price_col_index is None:
        print(f"⚠️ Column '{price_column}' not found in the Excel file.")
        return

    # Define red fill style
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Apply formatting to rows where the price is 'SOLD'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=price_col_index, max_col=price_col_index):
        if row[0].value == "SOLD":
            for cell in ws[row[0].row]:
                cell.fill = red_fill

    wb.save(filename)
    wb.close()
    print("✅ Red highlighting applied for 'SOLD' vehicles.")

def remove_unavailable_urls(input_filename, unavailable_urls):
    """Removes unavailable URLs from the input file."""
    if not os.path.exists(input_filename):
        print("⚠️ Input file does not exist.")
        return

    df_input = pd.read_excel(input_filename, dtype=str)

    if "URL" not in df_input.columns:
        print("⚠️ 'URL' column not found in input file.")
        return

    initial_count = len(df_input)

    # Remove rows where the URL is in the unavailable list
    df_filtered = df_input[~df_input["URL"].isin(unavailable_urls)]

    removed_count = initial_count - len(df_filtered)
    if removed_count > 0:
        df_filtered.to_excel(input_filename, index=False)
        print(f"✅ Removed {removed_count} unavailable URLs from {input_filename}")
    else:
        print("✅ No URLs needed to be removed.")

if __name__ == '__main__':
    # Read the input Excel file containing the URLs.
    df_input = pd.read_excel(INPUT_EXCEL_FILE)
    urls = df_input["URL"].tolist()
    autotrader_scraper_selenium(urls)
