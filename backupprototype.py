import os
import re
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Output and Input Excel files
OUTPUT_EXCEL_FILE = r"E:\Coding Projects\autotrader_data.xlsx"
INPUT_EXCEL_FILE = r"E:\Coding Projects\urls_input.xlsx"

def clean_mileage(mileage_text):
    """Extracts full mileage as a clean number, removing 'miles' and other text."""
    mileage_text = mileage_text.replace(",", "")  # Remove commas
    mileage_numbers = re.search(r'\d+', mileage_text)  # Extract the full mileage
    return mileage_numbers.group(0) if mileage_numbers else "Mileage not found"

def clean_registration_year(reg_text):
    """Extracts the full registration year and suffix, e.g., '2024 (74 reg)'."""
    reg_match = re.search(r'\d{4} \(\d{2} reg\)', reg_text)
    return reg_match.group(0) if reg_match else "Year not found"

def autotrader_scraper_selenium(urls):
    # Set up Selenium WebDriver
    options = Options()
    options.add_argument("_tt_enable_cookie=1") 
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--window-size=1920,1080")

    driver_service = Service(r"C:\Windows\chromedriver-win64\chromedriver.exe")  
    driver = webdriver.Chrome(service=driver_service, options=options)

    results = []
    today_date = datetime.today().strftime('%d-%m-%Y')  
    unavailable_urls = []

    for url in urls:
        try:
            print(f"Scraping: {url}")
            driver.get(url)

            # Check if the advert is no longer available
            try:
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'The advert you are looking for is no longer available')]"))
                )
                unavailable_urls.append(url)
                price_text = "SOLD"  
                mileage_text = "N/A"
                registration_year = "N/A"
            except:
                # Extract price
                try:
                    price_element = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "//h2[@data-testid='advert-price']"))
                    )
                    price_text = price_element.text.strip()
                except:
                    price_text = "Price not found"

                # Extract mileage
                try:
                    mileage_element = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "//li[contains(@class, 'at__sc-1n64n0d-9') and contains(@class, 'at__sc-1ebejir-1')]"))
                    )
                    mileage_text = clean_mileage(mileage_element.text.strip())  # Clean mileage
                except:
                    mileage_text = "Mileage not found"

                # Extract registration year
                try:
                    reg_elements = WebDriverWait(driver, 10).until(
                        EC.presence_of_all_elements_located((By.XPATH, "//ul[contains(@class, 'at__sc-1ebejir-0')]/li[contains(@class, 'at__sc-1n64n0d-9')]"))
                    )
                    if len(reg_elements) > 1:  # The second <li> usually contains the registration year
                        registration_year = clean_registration_year(reg_elements[1].text.strip())
                    else:
                        registration_year = "Year not found"
                except:
                    registration_year = "Year not found"

                # Extract make
                try:
                    make_element = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "//h1[@data-testid='advert-title']"))
                    )
                    make_text = make_element.text.strip()
                except:
                    make_text = "Make not found"

            # Store the result
            results.append({
                "URL": url,
                "Make": make_text,
                "Registration Year": registration_year,
                "Mileage": mileage_text,
                today_date: price_text
            })

        except Exception as e:
            print(f"Error scraping {url}: {e}")
            results.append({
                "URL": url,
                "Make": "Error fetching make",
                "Registration Year": "Error fetching year",
                "Mileage": "Error fetching mileage",
                today_date: "Error fetching price"
            })

    driver.quit()

    # Convert results to DataFrame
    df_new = pd.DataFrame(results)

    # Load existing data or create a new one
    if os.path.exists(OUTPUT_EXCEL_FILE):
        df_existing = pd.read_excel(OUTPUT_EXCEL_FILE, dtype=str)

        # Ensure only one "Mileage" column
        if "Mileage" in df_existing.columns:
            df_existing = df_existing.drop(columns=["Mileage"])
        
        # Ensure only one "Registration Year" column
        if "Registration Year" in df_existing.columns:
            df_existing = df_existing.drop(columns=["Registration Year"])

        df_combined = df_existing.merge(df_new, on=["URL", "Make"], how="outer")
    else:
        df_combined = df_new

    # Ensure required columns exist before reordering
    required_columns = ["URL", "Make", "Registration Year", "Mileage"]
    for col in required_columns:
        if col not in df_combined.columns:
            df_combined[col] = "N/A"  # Add missing columns with default "N/A" values

    # Reorder columns: "Mileage" remains next to "Registration Year"
    columns_order = ["URL", "Make", "Registration Year", "Mileage"] + [
        col for col in df_combined.columns if col not in ["URL", "Make", "Registration Year", "Mileage"]
    ]
    df_combined = df_combined[columns_order]

    # Save updated data to Excel
    df_combined.to_excel(OUTPUT_EXCEL_FILE, index=False)

    # Apply conditional formatting for SOLD listings
    apply_red_highlight(OUTPUT_EXCEL_FILE, today_date)

    # Remove unavailable URLs from input list
    if unavailable_urls:
        remove_unavailable_urls(INPUT_EXCEL_FILE, unavailable_urls)

    print(f"\n✅ Data saved to {OUTPUT_EXCEL_FILE}")

def apply_red_highlight(filename, price_column):
    """Applies red highlight to rows where the price is 'SOLD'."""
    wb = load_workbook(filename)
    ws = wb.active

    # Identify the column index for today's price
    price_col_index = None
    for col_index, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        if col[0].value == price_column:
            price_col_index = col_index
            break

    if price_col_index is None:
        print(f"⚠️ Column '{price_column}' not found in the Excel file.")
        return

    # Define red fill style
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Apply formatting to rows where the price is 'SOLD'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=price_col_index, max_col=price_col_index):
        if row[0].value == "SOLD":
            for cell in ws[row[0].row]:
                cell.fill = red_fill

    wb.save(filename)
    wb.close()
    print("✅ Red highlighting applied for 'SOLD' vehicles.")

def remove_unavailable_urls(input_filename, unavailable_urls):
    """Removes unavailable URLs from the input file."""
    if not os.path.exists(input_filename):
        print("⚠️ Input file does not exist.")
        return

    df_input = pd.read_excel(input_filename, dtype=str)

    if "URL" not in df_input.columns:
        print("⚠️ 'URL' column not found in input file.")
        return

    df_filtered = df_input[~df_input["URL"].isin(unavailable_urls)]
    df_filtered.to_excel(input_filename, index=False)

if __name__ == '__main__':
    df_input = pd.read_excel(INPUT_EXCEL_FILE)
    urls = df_input["URL"].tolist()
    autotrader_scraper_selenium(urls)
